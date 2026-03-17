import io
from datetime import datetime
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import Order, OrderItem


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ("product", "quantity", "price")


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "status", "total", "created_at")
    list_filter = ("status",)
    search_fields = ("user__telegram_id", "phone", "full_name")
    inlines = [OrderItemInline]
    readonly_fields = ("created_at", "updated_at")
    actions = ["export_paid_to_excel"]

    @admin.action(description="Экспорт оплаченных заказов в Excel")
    def export_paid_to_excel(self, request, queryset):
        qs = queryset.filter(status=Order.Status.PAID).select_related("user").prefetch_related("items__product").order_by("id")
        if not qs.exists():
            self.message_user(request, "Нет оплаченных заказов в выбранном наборе.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Оплаченные заказы"
        headers = ["ID", "Telegram ID", "ФИО", "Телефон", "Адрес", "Сумма", "Дата", "Позиции"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h).font = Font(bold=True)

        for row_idx, order in enumerate(qs, 2):
            items_str = "; ".join(
                f"{item.product.name} x{item.quantity}" for item in order.items.select_related("product")
            )
            ws.cell(row=row_idx, column=1, value=order.id)
            ws.cell(row=row_idx, column=2, value=order.user.telegram_id)
            ws.cell(row=row_idx, column=3, value=order.full_name)
            ws.cell(row=row_idx, column=4, value=order.phone)
            ws.cell(row=row_idx, column=5, value=order.address)
            ws.cell(row=row_idx, column=6, value=float(order.total))
            ws.cell(row=row_idx, column=7, value=order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "")
            ws.cell(row=row_idx, column=8, value=items_str)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"orders_paid_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={filename}"
        return response
